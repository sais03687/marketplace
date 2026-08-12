"""
MCP Python Sandbox — lightweight code execution + document parsing sidecar.

Runs inside a Docker container on the same internal network as the agent.
No external network access — only reachable by the agent container.

Endpoints:
  POST /mcp/tools/list  — list available tools
  POST /mcp/tools/call  — execute a tool
  GET  /health          — health check
"""

import base64
import io
import json
import os
import resource
import signal
import subprocess
import sys
import tempfile
from pathlib import Path

from fastapi import FastAPI, HTTPException
from pydantic import BaseModel

app = FastAPI(title="MCP Python Sandbox", version="1.0.0")

# ─── Tool Definitions ───────────────────────────────────────────────────────

TOOLS = [
    {
        "name": "execute_python",
        # This string is the only thing the agent knows about the sandbox — it is
        # written into MCP_TOOLS.md at startup and read by the model. It was
        # listing five of the nine installed libraries, so pdfplumber and
        # python-docx were invisible even though the agent is sold on "extracts
        # data from PDFs and Word docs". It also only said what *was* present,
        # which does not stop a model reaching for what is not: one run wasted an
        # iteration on engine="xlsxwriter" before recovering.
        #
        # So: name everything, say plainly that the list is closed, and say why
        # installing more will not work.
        "description": (
            "Execute Python code for data analysis, visualisation, computation and "
            "document parsing.\n"
            "\n"
            "Available libraries — this list is complete, nothing else is installed:\n"
            "  data      pandas, numpy\n"
            "  charts    matplotlib, seaborn\n"
            "  excel     openpyxl, xlsxwriter (writing), xlrd (legacy .xls)\n"
            "  documents pdfplumber (PDF), python-docx (Word)\n"
            "  text      tabulate, charset-normalizer\n"
            "\n"
            "There is no network access, so pip install will hang until the timeout "
            "rather than fetching anything. If you need something that is not listed, "
            "solve it with what is here or say you cannot.\n"
            "\n"
            "Limits: 30s per execution, 256MB memory. Load large inputs in chunks — "
            "exceeding the memory limit kills the process without a traceback.\n"
            "\n"
            "Files written to /tmp/output/ come back as a file_id, not as content. The "
            "platform holds the bytes; pass that file_id straight to drive_upload. You "
            "will not see the file contents and do not need to.\n"
            "\n"
            "This is the only way to get a file out of the sandbox. Uploads accept a "
            "file_id and reject file content, so never paste base64 and never invent an "
            "id — if you have not written a file, you do not have one to upload.\n"
            "\n"
            "Files come IN the same way. When an email arrives with an attachment you "
            "are given a handle for it; list those handles in input_files and the "
            "platform writes each one to /tmp/input/<filename> before your code runs. "
            "Read them with pandas, openpyxl, pdfplumber — whatever suits the format. "
            "Never paste file content into code: you were given a handle precisely "
            "because the file is too big to retype, and a handle is the only way in."
        ),
        "inputSchema": {
            "type": "object",
            "properties": {
                "code": {
                    "type": "string",
                    "description": "Python code to execute. Print results to stdout. Save files to /tmp/output/ to return them.",
                },
                "input_files": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": (
                        "Handles of files to place in /tmp/input/ before the code runs — "
                        "the ones you were given for email attachments. Names are kept, "
                        "so an attached orders.csv is at /tmp/input/orders.csv."
                    ),
                },
            },
            "required": ["code"],
        },
    },
    {
        "name": "parse_pdf",
        "description": "Extract text and tables from a PDF file. Returns structured text content.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "file_content_base64": {
                    "type": "string",
                    "description": "Base64-encoded PDF file content",
                },
            },
            "required": ["file_content_base64"],
        },
    },
    {
        "name": "parse_docx",
        "description": "Extract text from a Word document (.docx).",
        "inputSchema": {
            "type": "object",
            "properties": {
                "file_content_base64": {
                    "type": "string",
                    "description": "Base64-encoded .docx file content",
                },
            },
            "required": ["file_content_base64"],
        },
    },
    {
        "name": "parse_xlsx",
        "description": "Extract sheet data from an Excel file as JSON.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "file_content_base64": {
                    "type": "string",
                    "description": "Base64-encoded .xlsx file content",
                },
            },
            "required": ["file_content_base64"],
        },
    },
]


# ─── Models ──────────────────────────────────────────────────────────────────

class ToolCallRequest(BaseModel):
    name: str
    arguments: dict


# ─── Endpoints ───────────────────────────────────────────────────────────────

@app.get("/health")
async def health():
    return {"ok": True}


@app.post("/mcp/tools/list")
async def list_tools():
    return {"tools": TOOLS}


@app.post("/mcp/tools/call")
async def call_tool(req: ToolCallRequest):
    handler = TOOL_HANDLERS.get(req.name)
    if not handler:
        raise HTTPException(status_code=404, detail=f"Unknown tool: {req.name}")
    try:
        result = handler(req.arguments)
        return {"result": result}
    except Exception as e:
        return {"error": str(e)}


# ─── Tool Implementations ───────────────────────────────────────────────────

def _stage_input_files(entries: list) -> tuple[list[str], list[str]]:
    """Write inbound files into /tmp/input/ and return (staged names, errors).

    The adapter has already turned the handles the model passed into
    {name, content_base64} pairs — the model never sees or types base64, in
    either direction. Anything malformed is reported rather than skipped
    silently: code written against a file that is not there fails with a
    FileNotFoundError that says nothing about why.
    """
    in_dir = Path("/tmp/input")
    staged: list[str] = []
    errors: list[str] = []

    # Clear between runs so one execution cannot read another's inputs.
    if in_dir.exists():
        for old in in_dir.iterdir():
            if old.is_file():
                old.unlink()
    in_dir.mkdir(parents=True, exist_ok=True)

    for entry in entries:
        if not isinstance(entry, dict):
            errors.append(f"unusable input_files entry: {type(entry).__name__}")
            continue
        raw_name = str(entry.get("name") or "input")
        # basename only — an attachment called ../../etc/passwd must land in
        # /tmp/input, not where it asked to.
        name = os.path.basename(raw_name).replace("\\", "_") or "input"
        try:
            blob = base64.b64decode(entry.get("content_base64") or "")
        except Exception as exc:
            errors.append(f"{name}: could not be decoded ({exc})")
            continue
        try:
            (in_dir / name).write_bytes(blob)
            staged.append(name)
        except Exception as exc:
            errors.append(f"{name}: could not be written ({exc})")

    return staged, errors


def _execute_python(args: dict) -> dict:
    """Run Python code in a subprocess with resource limits."""
    code = args.get("code", "")
    if not code.strip():
        return {"stdout": "", "stderr": "No code provided", "files": []}

    staged, stage_errors = _stage_input_files(args.get("input_files") or [])
    if staged:
        print(f"[sandbox] Staged {len(staged)} input file(s): {', '.join(staged)}", flush=True)

    # Create output directory for generated files
    output_dir = tempfile.mkdtemp(prefix="mcp_output_")
    output_path = Path(output_dir)

    # Wrap code to set up output dir
    wrapper = f"""
import os, sys
os.makedirs('/tmp/output', exist_ok=True)
os.makedirs('/tmp/input', exist_ok=True)
os.chdir('/tmp')
{code}
"""

    try:
        result = subprocess.run(
            [sys.executable, "-c", wrapper],
            capture_output=True,
            text=True,
            timeout=30,
            env={
                "PATH": os.environ.get("PATH", "/usr/local/bin:/usr/bin:/bin"),
                "HOME": "/tmp",
                "MPLBACKEND": "Agg",  # matplotlib non-interactive backend
            },
        )

        # Collect any files written to /tmp/output/
        files = []
        out_dir = Path("/tmp/output")
        if out_dir.exists():
            for f in out_dir.iterdir():
                if f.is_file() and f.stat().st_size < 10_000_000:  # 10MB max per file
                    files.append({
                        "name": f.name,
                        "base64_content": base64.b64encode(f.read_bytes()).decode(),
                    })
                    f.unlink()  # clean up

        # A file that failed to stage is the likely cause of whatever the code
        # then did wrong, so say so ahead of the traceback rather than leaving
        # a bare FileNotFoundError to be puzzled over.
        stderr = result.stderr[:10000]
        if stage_errors:
            stderr = "Input files that could not be staged:\n" + "\n".join(
                f"  {e}" for e in stage_errors
            ) + ("\n\n" + stderr if stderr else "")

        return {
            "stdout": result.stdout[:50000],  # cap output at 50KB
            "stderr": stderr,
            "returncode": result.returncode,
            "files": files,
            **({"input_files_staged": staged} if staged else {}),
        }
    except subprocess.TimeoutExpired:
        return {
            "stdout": "",
            "stderr": "Execution timed out after 30 seconds",
            "returncode": -1,
            "files": [],
        }
    except Exception as e:
        return {
            "stdout": "",
            "stderr": f"Execution error: {str(e)}",
            "returncode": -1,
            "files": [],
        }


def _parse_pdf(args: dict) -> dict:
    """Extract text and tables from a PDF."""
    try:
        import pdfplumber
    except ImportError:
        return {"error": "pdfplumber not installed"}

    raw = base64.b64decode(args.get("file_content_base64", ""))
    text_parts = []
    tables = []

    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        for i, page in enumerate(pdf.pages):
            page_text = page.extract_text() or ""
            if page_text.strip():
                text_parts.append(f"--- Page {i + 1} ---\n{page_text}")
            page_tables = page.extract_tables()
            for table in page_tables:
                tables.append({"page": i + 1, "data": table})

    return {
        "text": "\n\n".join(text_parts),
        "tables": tables,
        "page_count": len(text_parts),
    }


def _parse_docx(args: dict) -> dict:
    """Extract text from a Word document."""
    try:
        import docx
    except ImportError:
        return {"error": "python-docx not installed"}

    raw = base64.b64decode(args.get("file_content_base64", ""))
    doc = docx.Document(io.BytesIO(raw))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]

    return {
        "text": "\n\n".join(paragraphs),
        "paragraph_count": len(paragraphs),
    }


def _parse_xlsx(args: dict) -> dict:
    """Extract sheet data from an Excel file."""
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed"}

    raw = base64.b64decode(args.get("file_content_base64", ""))
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    sheets = {}

    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        sheets[name] = rows

    return {"sheets": sheets}


TOOL_HANDLERS = {
    "execute_python": _execute_python,
    "parse_pdf": _parse_pdf,
    "parse_docx": _parse_docx,
    "parse_xlsx": _parse_xlsx,
}
